'''
读取excel分册页数信息，分组后，使用pypdf2逐页读取另存为，excel第一列为作品号，一个作品号存放在一个文件夹
一本书跨册的，最终合并到一个文件夹中
src = 原始pdf存放位置
dst = 拆页后文件存放位置
excel =拆页excel文件路径

'''
import os,io
from itertools import groupby
from PyPDF2 import PdfFileReader,PdfFileWriter
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook


class Book :
    def __init__(book,nameid,name,booknum,startpage,endpage):
        book.nameid = nameid     #作品序号
        book.name = name       #书名
        book.booknum = booknum     #所在册数
        book.startpage =startpage   #开始页
        book.endpage = endpage    #结束页


bookls =[]

src = r"F:\已拆完（pdf\aaaa"
dst = r"F:\lypdf"
excel = xlrd.open_workbook(r"F:\已拆完.xls")
table = excel.sheets()[0]
nrows = table.nrows  #行
nclos = table.ncols   #列

for i in range(0,nrows):
    nameid = table.cell(i,0).value
    name = table.cell(i,1).value
    booknum = table.cell(i,2).value
    startpage = table.cell(i,3).value
    endpage = table.cell(i,4).value
    book = Book(nameid,name,booknum,startpage,endpage)
    bookls.append(book)

bkls = sorted(bookls,key = lambda x:x.nameid)  #按照Book对象的nameid属性排序,然后用自带的groupby分组
for key,group in groupby(bkls,lambda x:x.nameid):
    gl =list(group)  #分组后长度，1可以单独拆出，大于1时，多册合并为一册
    glen = len(gl)
    if glen == 1:
        
        bkid = "{:0>3d}".format(int(gl[0].nameid))
        num = "{:0>3}".format(int(gl[0].booknum))
        starpage = int(gl[0].startpage)
        endpage = int(gl[0].endpage)
        name = bkid +"-"+gl[0].name

        srcpath = src +"\\"+ num +"\\"+num +"-ZW.pdf"
        savepath = dst +"\\"+name
        if not os.path.exists(savepath):
            os.mkdir(savepath)
        
        srcbook = PdfFileReader(open(srcpath,'rb'))
        
        
        for  i in range(starpage-1,endpage):
            pdfw =PdfFileWriter()
            page = srcbook.getPage(i)
            pname = os.path.join(savepath,"{:0>7d}".format(i)+".pdf")
            pdfw.addPage(page)
            pdfw.write(open(pname,'wb'))
        
        print("拆分完毕"+savepath)
        
            
    if glen >1:
        
        bkid = "{:0>3d}".format(int(gl[0].nameid))
        num = "{:0>3}".format(int(gl[0].booknum))
        starpage = int(gl[0].startpage)
        endpage = int(gl[0].endpage)
        name = bkid +"-"+gl[0].name

        srcpath = src +"\\"+ num +"\\"+num +"-ZW.PDF"
        savepath = dst +"\\"+name
        if not os.path.exists(savepath):
            os.mkdir(savepath)
        sumnum = 1      #总页数计数
        
        
        for bk in range(0,glen):
            num = "{:0>3}".format(int(gl[bk].booknum))
            starpage = int(gl[bk].startpage)
            endpage = int(gl[bk].endpage)
            srcpath = src +"\\"+ num +"\\"+num +"-ZW.PDF"
            srcbook = PdfFileReader(open(srcpath,'rb'))
            for  i in range(starpage-1,endpage):
                pdfw =PdfFileWriter()
                sumnum = 1+sumnum
                page = srcbook.getPage(i)
                pname = os.path.join(savepath,"{:0>7d}".format(sumnum)+".pdf")
                pdfw.addPage(page)
                pdfw.write(open(pname,'wb'))
        print("拆分完毕"+savepath)
 
            
