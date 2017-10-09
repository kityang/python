#!python3
'''
xlrd无法读取xlsx,表格超过65536行，需要使用openpyxl
读取txt中的关键词，然后遍历单元格，匹配到就输出到rst文件中
'''
import os
from openpyxl import Workbook
from openpyxl import load_workbook


rst = open(r"C:\Users\ygl\Desktop\result.txt",'w')

txtlist = []
with open(r"C:\Users\ygl\Desktop\关键词.txt",'r') as f:
    for line in f.readlines():
       txtlist.append(line.strip()) 




print ("正在读取excel...")
workbook = load_workbook(r"C:\Users\ygl\Desktop\总表.xlsx")
print ("excel读取完毕")
workbook.active
sheetname = workbook.get_sheet_names()[0]
sheet = workbook[sheetname]

row = sheet.max_row
clo = sheet.max_column

for i in range(1,row +1):
    for j in range(2,clo +1):
        cellv = sheet.cell(row =i,column =j).value
        if isinstance(cellv,str):
            for k in txtlist:
                if cellv.find(k) !=-1:
                    c1 = sheet.cell(row =i,column =1).value
                    c2 = sheet.cell(row =i,column =2).value
                    rst.write(c1+" "+c2+"\n")
                    print(c2)
                    break
                    
rst.close()
print("OK")

        
        
